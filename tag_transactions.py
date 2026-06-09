#!/usr/bin/env python3
"""
Tag output CSVs with categories and write to a single XLSX file.
Each account becomes one sheet. Run after extract_transactions.py.

Output: output/tagged.xlsx
"""

import csv, glob, json, re, collections
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font

ROOT = Path(__file__).parent

# ── Load rules ──────────────────────────────────────────────────────────────
with open(ROOT / 'report/category_rules.json', encoding='utf-8') as f:
    rules_data = json.load(f)

rules           = [(r['category'], [kw.upper() for kw in r['keywords']]) for r in rules_data['rules']]
manual_review   = [kw.upper() for kw in rules_data['manual_review']]
broad_merchants = [kw.upper() for kw in rules_data['broad_merchants']]

# ── Load all transactions ───────────────────────────────────────────────────
CSV_FIELDS = ['date', 'particulars', 'debit', 'credit', 'balance', 'pdf_file']

account_txns: dict[str, list[dict]] = {}

for f in sorted(glob.glob(str(ROOT / 'output/Transactions_*.csv'))):
    account = Path(f).stem.split('_', 1)[1]
    rows = []
    with open(f, encoding='utf-8') as fh:
        for line in fh:
            if not line.startswith('#'):
                break
        reader = csv.DictReader(fh, fieldnames=CSV_FIELDS)
        for row in reader:
            if row['date'] == 'date':
                continue
            rows.append({k: (v or '').strip() for k, v in row.items()})
    account_txns[account] = rows
    print(f'Loaded {len(rows):>5} rows  [{account}]')

# ── Internal transfer detection ─────────────────────────────────────────────
def _norm(p: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', p.upper())

by_key: dict = collections.defaultdict(list)
for account, rows in account_txns.items():
    for row in rows:
        amount = row['debit'] or row['credit']
        if amount and row['particulars']:
            by_key[(row['date'], amount, _norm(row['particulars']))].append((account, row))

internal_ids: set[int] = set()
for entries in by_key.values():
    if len({acc for acc, _ in entries}) > 1:
        for _, row in entries:
            internal_ids.add(id(row))

# ── Categorise ──────────────────────────────────────────────────────────────
def categorise(row: dict) -> tuple[str, bool]:
    """Return (category, needs_review)."""
    if id(row) in internal_ids:
        return 'Internal Transfer', False

    pu = row['particulars'].upper()

    for kw in manual_review:
        if kw in pu:
            return '', True

    cat = ''
    for rule_cat, kws in rules:
        if any(kw in pu for kw in kws):
            cat = rule_cat
            break

    needs_review = not cat or any(kw in pu for kw in broad_merchants)
    return cat, needs_review

# ── Write XLSX ───────────────────────────────────────────────────────────────
OUT_COLS     = CSV_FIELDS + ['category', 'needs_review']
HEADER_FONT  = Font(bold=True)
REVIEW_FILL  = PatternFill(start_color='FFFFD700', end_color='FFFFD700', fill_type='solid')

wb = openpyxl.Workbook()
wb.remove(wb.active)

totals = collections.Counter()

for account, rows in sorted(account_txns.items()):
    ws = wb.create_sheet(title=account)

    ws.append(OUT_COLS)
    for cell in ws[1]:
        cell.font = HEADER_FONT

    for row in rows:
        cat, needs_review = categorise(row)
        totals['total'] += 1
        if cat == 'Internal Transfer':
            totals['internal'] += 1
        elif needs_review:
            totals['review'] += 1
        else:
            totals['tagged'] += 1

        ws.append([
            row['date'], row['particulars'], row['debit'], row['credit'],
            row['balance'], row['pdf_file'],
            cat,
            'YES' if needs_review else '',
        ])

        if needs_review:
            for cell in ws[ws.max_row]:
                cell.fill = REVIEW_FILL

output_path = ROOT / 'output/tagged.xlsx'
wb.save(output_path)

print(f'\nWritten: {output_path}')
print(f'  {totals["total"]:>5}  total')
print(f'  {totals["internal"]:>5}  internal transfers (auto)')
print(f'  {totals["tagged"]:>5}  tagged')
print(f'  {totals["review"]:>5}  needs review (highlighted)')
